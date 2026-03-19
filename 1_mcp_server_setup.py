import asyncio
import os
import fitz  # PyMuPDF
import openpyxl
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Resource, TextContent, ReadResourceResult

# ── Configuration ──────────────────────────────────────────────
PDF_FOLDER = "data/blueprints/"
EXCEL_PATH = "data/master_book.xlsx"

# ── Initialize MCP Server ──────────────────────────────────────
app = Server("engineering-doc-server")


@app.list_resources()
async def list_resources() -> list[Resource]:
    """
    Exposes all PDF blueprints and the Excel master book
    as standardized queryable resources to the LLM.
    """
    resources = []

    # Register each PDF blueprint as a resource
    if os.path.exists(PDF_FOLDER):
        for filename in os.listdir(PDF_FOLDER):
            if filename.endswith(".pdf"):
                resources.append(Resource(
                    uri=f"file://blueprints/{filename}",
                    name=filename,
                    description=f"Engineering blueprint: {filename}",
                    mimeType="application/pdf"
                ))

    # Register the Excel master book as a resource
    if os.path.exists(EXCEL_PATH):
        resources.append(Resource(
            uri="file://excel/master_book.xlsx",
            name="Master Engineering Book",
            description="Excel repository of historical engineering records",
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ))

    return resources


@app.read_resource()
async def read_resource(uri: str) -> ReadResourceResult:
    """
    Reads and returns the content of a requested resource.
    Handles both PDF blueprints and Excel master books.
    """

    # ── Handle PDF blueprints ──────────────────────────────────
    if uri.startswith("file://blueprints/"):
        filename = uri.replace("file://blueprints/", "")
        filepath = os.path.join(PDF_FOLDER, filename)

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Blueprint not found: {filename}")

        # Extract text from all pages using PyMuPDF
        doc = fitz.open(filepath)
        full_text = ""
        for page_num, page in enumerate(doc):
            full_text += f"\n--- Page {page_num + 1} ---\n"
            full_text += page.get_text()
        doc.close()

        return ReadResourceResult(
            contents=[TextContent(
                type="text",
                text=full_text
            )]
        )

    # ── Handle Excel master book ───────────────────────────────
    elif uri == "file://excel/master_book.xlsx":
        if not os.path.exists(EXCEL_PATH):
            raise FileNotFoundError("Excel master book not found")

        wb = openpyxl.load_workbook(EXCEL_PATH)
        full_text = ""

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            full_text += f"\n=== Sheet: {sheet_name} ===\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(
                    str(cell) for cell in row if cell is not None
                )
                if row_text.strip():
                    full_text += row_text + "\n"

        return ReadResourceResult(
            contents=[TextContent(
                type="text",
                text=full_text
            )]
        )

    else:
        raise ValueError(f"Unknown resource URI: {uri}")


# ── Entry Point ────────────────────────────────────────────────
if __name__ == "__main__":
    asyncio.run(stdio_server(app))
